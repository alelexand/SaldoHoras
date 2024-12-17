import openpyxl
import sqlite3


class Consulta:
    def __init__(self):
        self.banco = sqlite3.connect(fr"CaminhoBancoDeDados")
        self.cursor = self.banco.cursor()
        self.wb = openpyxl.load_workbook(fr'CaminhoExcelFormatadoParaReceberOsDados')
        self.ws = self.wb.active
        self.lista_nomes_db = []
        self.lista_final = []

    def horas(self):

        lista_hora = []
        row = 2
        self.cursor.execute(f'SELECT nome FROM registro')
        nome1 = self.cursor.fetchall()
        for nome in nome1:
            self.lista_nomes_db.append(nome[0])

        while row <= self.ws.max_row:
            nome_pessoa = self.ws[f'A{row}'].value

            if nome_pessoa in self.lista_nomes_db:
                # print(nome_pessoa + f' LINHA: {row}')

                for linha in range(row, self.ws.max_row + 1):
                    nome_saldo = self.ws[f'H{linha}'].value

                    if nome_saldo == 'Saldo do Período:':
                        saldo = self.ws[f'N{linha}'].value
                        # print(f"SALDO DE HORAS É {saldo}")
                        break
                lista_hora = [nome_pessoa, saldo]
                self.lista_final.append(lista_hora)
                row = linha
            else:
                row += 1
        # print(self.lista_final)
        return self.lista_final


